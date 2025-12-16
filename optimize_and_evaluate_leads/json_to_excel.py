#!/usr/bin/env python3
from pathlib import Path
import json
from typing import Any, Dict, List
import pandas as pd
import re

FALLBACK_JSON = Path("outputs/lead_list_sorted.json")
FALLBACK_XLSX = Path("outputs/final_leads_list.xlsx")

# ============================================================
# Excel-safe sanitization (HARD GUARANTEE)
# ============================================================

_ILLEGAL_EXCEL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

def sanitize_for_excel(value: Any) -> Any:
    """
    Removes all characters that Excel / openpyxl cannot handle.
    Safe for strings, lists, dicts, numbers, None.
    """
    if isinstance(value, str):
        return _ILLEGAL_EXCEL_CHARS.sub("", value)
    if isinstance(value, list):
        return [sanitize_for_excel(v) for v in value]
    if isinstance(value, dict):
        return {k: sanitize_for_excel(v) for k, v in value.items()}
    return value


def leads_json_to_excel_preserve(
    input_path: Path,
    excel_path: Path,
) -> None:

    if excel_path is None:
        excel_path = FALLBACK_XLSX

    if input_path is None:
        input_path = FALLBACK_JSON

    excel_path = Path(excel_path)
    input_path = Path(input_path)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    excel_path.parent.mkdir(parents=True, exist_ok=True)

    with input_path.open("r", encoding="utf-8") as f:
        loaded = json.load(f)

    # ðŸ”’ sanitize immediately after load
    loaded = sanitize_for_excel(loaded)

    if isinstance(loaded, dict) and "leads" in loaded:
        items = loaded["leads"]
    elif isinstance(loaded, list):
        items = loaded
    else:
        items = [loaded]

    items = [it if isinstance(it, dict) else {} for it in items]

    max_src = 0
    for it in items:
        v = it.get("source_urls", [])
        if isinstance(v, list):
            max_src = max(max_src, len(v))

    rows = []
    all_keys = set()

    for it in items:
        row = dict(it)

        srcs = it.get("source_urls", [])
        if not isinstance(srcs, list):
            srcs = [] if srcs is None else [srcs]

        for i in range(max_src):
            col = f"source_url_{i+1}"
            row[col] = srcs[i] if i < len(srcs) else ""
            all_keys.add(col)

        for k in ("company", "website", "mail", "phone_number", "location", "description"):
            all_keys.add(k)
            if k not in row:
                row[k] = ""

        for k in it.keys():
            all_keys.add(k)

        # ðŸ”’ final row-level sanitation
        rows.append(sanitize_for_excel(row))

    preferred = [
        "company",
        "website",
        "mail",
        "phone_number",
        "location",
        "description",
    ]
    ordered = preferred + [k for k in sorted(all_keys) if k not in preferred]

    df = pd.DataFrame(rows, columns=ordered)

    # ðŸ”’ sanitize dataframe values one last time
    df = df.applymap(sanitize_for_excel)

    df.to_excel(excel_path, index=False, engine="openpyxl")

    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        wb = load_workbook(excel_path)
        ws = wb.active

        headers = {
            ws.cell(row=1, column=i).value: i
            for i in range(1, ws.max_column + 1)
        }

        def set_hyperlink(row_idx: int, col_idx: int):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, str):
                v = cell.value
                if v.startswith("http://") or v.startswith("https://"):
                    cell.hyperlink = v
                    cell.style = "Hyperlink"
            cell.number_format = "@"

        for r in range(2, ws.max_row + 1):
            if "website" in headers:
                set_hyperlink(r, headers["website"])

            if "phone_number" in headers:
                c = ws.cell(row=r, column=headers["phone_number"])
                if c.value is not None:
                    c.value = str(c.value)
                c.number_format = "@"

            if "mail" in headers:
                c = ws.cell(row=r, column=headers["mail"])
                if c.value is not None:
                    c.value = str(c.value)
                c.number_format = "@"

            for i in range(1, max_src + 1):
                colname = f"source_url_{i}"
                if colname in headers:
                    set_hyperlink(r, headers[colname])

        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_len = 0
            for cell in ws[col_letter]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(80, max(10, max_len + 2))

        wb.save(excel_path)
    except Exception:
        pass

    print(f"Final Excel written to {excel_path}")
